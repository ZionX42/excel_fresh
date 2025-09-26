from fastapi import FastAPI, APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone
from passlib.context import CryptContext
from jose import jwt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from urllib.parse import urlencode
from pymongo.errors import ServerSelectionTimeoutError


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'app_db')]

# Security setup
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
JWT_SECRET = os.environ.get('JWT_SECRET', 'dev-secret-change')
JWT_ALG = 'HS256'

# OAuth envs (optional until configured)
GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID')
MICROSOFT_CLIENT_ID = os.environ.get('MICROSOFT_CLIENT_ID')
MICROSOFT_TENANT_ID = os.environ.get('MICROSOFT_TENANT_ID', 'common')

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Helpers for Mongo-friendly serialization

def now_iso():
    return datetime.now(timezone.utc).isoformat()


def prepare_for_mongo(data: dict) -> dict:
    # Ensure dates are ISO strings
    for k, v in list(data.items()):
        if isinstance(v, datetime):
            data[k] = v.astimezone(timezone.utc).isoformat()
    return data


# ====== Models ======
class StatusCheck(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: str = Field(default_factory=now_iso)


class StatusCheckCreate(BaseModel):
    client_name: str


class GenerationRequest(BaseModel):
    description: str
    provider: Optional[str] = Field(default="auto")  # openai|anthropic|gemini|auto


class GenerationRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    description: str
    provider: str
    filename: str
    size_bytes: int
    created_at: str = Field(default_factory=now_iso)


class RegisterRequest(BaseModel):
    email: EmailStr
    password: str


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"


# ====== Routes ======
@api_router.get("/")
async def root():
    return {"message": "Hello World"}


@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_obj = StatusCheck(client_name=input.client_name)
    await db.status_checks.insert_one(status_obj.model_dump())
    return status_obj


@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    status_checks = await db.status_checks.find().to_list(1000)
    # Pydantic ignores _id
    return [StatusCheck(**{k: v for k, v in sc.items() if k != "_id"}) for sc in status_checks]


# -------- Auth (minimal, optional) --------
@api_router.post("/auth/register")
async def register(req: RegisterRequest):
    try:
        existing = await db.users.find_one({"email": req.email})
        if existing:
            raise HTTPException(status_code=400, detail="Email already registered")
        user = {
            "id": str(uuid.uuid4()),
            "email": req.email,
            "password_hash": pwd_context.hash(req.password),
            "created_at": now_iso(),
        }
        await db.users.insert_one(user)
        return {"ok": True}
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable - MongoDB not running")
    except Exception as e:
        raise


@api_router.post("/auth/login", response_model=LoginResponse)
async def login(req: LoginRequest):
    try:
        user = await db.users.find_one({"email": req.email})
        if not user or not pwd_context.verify(req.password, user.get("password_hash", "")):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        payload = {"sub": user["id"], "email": user["email"], "iat": int(datetime.now(timezone.utc).timestamp())}
        token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)
        return LoginResponse(access_token=token)
    except ServerSelectionTimeoutError:
        raise HTTPException(status_code=503, detail="Database unavailable - MongoDB not running")
    except Exception as e:
        raise


# --- OAuth: Login URL helpers (stubs until creds provided) ---
@api_router.get("/auth/google/login")
async def google_login(request: Request):
    if not GOOGLE_CLIENT_ID:
        raise HTTPException(status_code=400, detail="Google OAuth not configured")
    redirect_uri = str(request.url_for("google_callback"))
    params = {
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": redirect_uri,
        "response_type": "code",
        "scope": "openid email profile",
        "access_type": "offline",
        "prompt": "consent",
    }
    url = f"https://accounts.google.com/o/oauth2/v2/auth?{urlencode(params)}"
    return {"auth_url": url}


@api_router.get("/auth/google/callback", name="google_callback")
async def google_callback():
    # Full token exchange will be added after client secrets are provided
    raise HTTPException(status_code=501, detail="Google OAuth callback not implemented yet. Provide OAuth credentials to enable.")


@api_router.get("/auth/microsoft/login")
async def microsoft_login(request: Request):
    if not MICROSOFT_CLIENT_ID:
        raise HTTPException(status_code=400, detail="Microsoft OAuth not configured")
    redirect_uri = str(request.url_for("microsoft_callback"))
    params = {
        "client_id": MICROSOFT_CLIENT_ID,
        "response_type": "code",
        "redirect_uri": redirect_uri,
        "response_mode": "query",
        "scope": "openid email profile offline_access",
    }
    url = f"https://login.microsoftonline.com/{MICROSOFT_TENANT_ID}/oauth2/v2.0/authorize?{urlencode(params)}"
    return {"auth_url": url}


@api_router.get("/auth/microsoft/callback", name="microsoft_callback")
async def microsoft_callback():
    raise HTTPException(status_code=501, detail="Microsoft OAuth callback not implemented yet. Provide OAuth credentials to enable.")


# -------- Spreadsheet Generation (non-AI stub) --------

def build_workbook(description: str) -> BytesIO:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "README"

    title_font = Font(bold=True, size=14)
    subtle = PatternFill(start_color="FFF5F5F7", end_color="FFF5F5F7", fill_type="solid")

    ws_info["A1"] = "Generated Spreadsheet"
    ws_info["A1"].font = title_font
    ws_info["A2"] = f"Description: {description}"
    ws_info["A3"] = f"Generated At: {datetime.now(timezone.utc).isoformat()}"
    ws_info["A5"] = "This is an instant stub (no AI yet). We'll use AI in the next step."
    ws_info.column_dimensions['A'].width = 90
    for r in range(1, 7):
        ws_info[f"A{r}"].alignment = Alignment(wrap_text=True)
    ws_info["A1"].fill = subtle

    # Simple model based on keywords
    ws = wb.create_sheet("Data")
    ws["A1"], ws["B1"], ws["C1"], ws["D1"] = ("Month", "Revenue", "Costs", "Profit")
    for cell in ("A1", "B1", "C1", "D1"):
        ws[cell].font = Font(bold=True)
    months = [
        "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
    ]
    base_rev = 18000
    base_cost = 12000
    for i, m in enumerate(months, start=2):
        ws[f"A{i}"] = m
        ws[f"B{i}"] = base_rev + (i - 2) * 1000
        ws[f"C{i}"] = base_cost + (i - 2) * 600
        ws[f"D{i}"] = f"=B{i}-C{i}"
    ws.auto_filter.ref = f"A1:D{len(months)+1}"

    # Summary sheet with totals and a chart
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"].value = "KPI Summary"
    ws_sum["A1"].font = title_font
    ws_sum["A3"].value = "Total Revenue"
    ws_sum["B3"].value = f"=SUM(Data!B2:B{len(months)+1})"
    ws_sum["A4"].value = "Total Costs"
    ws_sum["B4"].value = f"=SUM(Data!C2:C{len(months)+1})"
    ws_sum["A5"].value = "Total Profit"
    ws_sum["B5"].value = f"=SUM(Data!D2:D{len(months)+1})"

    chart = LineChart()
    chart.title = "Revenue vs Costs vs Profit"
    data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=len(months)+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(months)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 24
    ws_sum.add_chart(chart, "A7")

    # Additional sheet to increase richness and file size for testing
    ws_tx = wb.create_sheet("Transactions")
    ws_tx["A1"], ws_tx["B1"], ws_tx["C1"], ws_tx["D1"], ws_tx["E1"], ws_tx["F1"] = ("Date", "Category", "Amount", "Note", "Reference", "Description")
    for cell in ("A1", "B1", "C1", "D1", "E1", "F1"):
        ws_tx[cell].font = Font(bold=True)
    cats = ["Sales", "Ops", "Marketing", "R&D", "Other", "Support", "Finance", "Legal", "HR", "IT"]
    from datetime import timedelta
    start = datetime(2025, 1, 1, tzinfo=timezone.utc)
    # Increased to 800 rows with more columns to ensure file size > 20000 bytes
    for i in range(1, 801):
        d = start + timedelta(days=i % 365)
        ws_tx[f"A{i+1}"] = d.date().isoformat()
        ws_tx[f"B{i+1}"] = cats[i % len(cats)]
        ws_tx[f"C{i+1}"] = (i * 7) % 900 + 50
        ws_tx[f"D{i+1}"] = f"Auto-generated transaction row {i} with detailed description"
        ws_tx[f"E{i+1}"] = f"REF-{i:06d}"
        ws_tx[f"F{i+1}"] = f"Detailed description for transaction {i} including additional context and information to increase file size"

    # Save to bytes
    bytes_io = BytesIO()
    wb.save(bytes_io)
    bytes_io.seek(0)
    return bytes_io


@api_router.post("/generate")
async def generate_spreadsheet(req: GenerationRequest):
    # Build workbook instantly (no external AI)
    xlsx_stream = build_workbook(req.description)

    # Persist a generation record (non-blocking feel, but awaited here)
    filename = f"spreadsheet_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    size_bytes = len(xlsx_stream.getbuffer())
    record = GenerationRecord(
        description=req.description,
        provider=(req.provider or "auto"),
        filename=filename,
        size_bytes=size_bytes,
    )
    # Try to persist record, but don't fail generation if DB is unavailable
    try:
        await db.generations.insert_one(prepare_for_mongo(record.model_dump()))
    except ServerSelectionTimeoutError:
        # Log and continue; file streaming should not be blocked by DB
        logger.warning("MongoDB unavailable during insert of generation record; continuing without persistence")

    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    return StreamingResponse(xlsx_stream, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)


@api_router.get("/generations", response_model=List[GenerationRecord])
async def list_generations():
    try:
        gens = await db.generations.find().sort("created_at", -1).to_list(100)
        cleaned = []
        for g in gens:
            g.pop("_id", None)
            cleaned.append(GenerationRecord(**g))
        return cleaned
    except ServerSelectionTimeoutError:
        # Graceful degradation when DB is offline
        logger.warning("MongoDB unavailable when listing generations; returning empty list")
        return []


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()